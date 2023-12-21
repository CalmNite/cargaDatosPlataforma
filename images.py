import json
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import base64
from io import BytesIO
import pandas as pd
import os

lista = []
file_names = [os.getenv("FILE_NAME") + str(i) + '.xlsx' for i in range(1, 23)]


for filename in file_names:
    workbook = load_workbook(filename)
    tf = pd.read_excel(filename)
    worksheet = workbook.active 
    image_loader = SheetImageLoader(worksheet)

    i = 1
    for row in worksheet.iter_rows():
        if row[0].value != None:
            if row[1].value == True:
                try:
                    image = image_loader.get(f'E{i}')
                    buffered = BytesIO()
                    image.save(buffered, format="JPEG")
                    img_str = base64.b64encode(buffered.getvalue())
                    img_base64 = bytes("data:image/jpeg;base64,", encoding='utf-8') + img_str
                    img_base64 = img_base64.decode("utf-8")
                    lista.append({"codigo": str(row[0].value).split(".")[0] , "imagen":  img_base64})
                except:
                    continue
        
        i+=1

with open('json_data.json', 'w') as outfile:
    json_string = json.dumps(lista)
    outfile.write(json_string)